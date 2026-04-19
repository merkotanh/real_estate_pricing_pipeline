import logging

logger = logging.getLogger(__name__)

import pandas as pd

def validate_merge(df):
    ratio = df['bundesland'].notna().mean()

    if ratio < 0.8:
        logging.warning(f"[bundesland] Low coverage: {ratio:.2f}")
    else:
        logging.info(f"[bundesland] Good coverage: {ratio:.2f}")

def validate_mean_columns(df):
    checks = {
        'price_euro': df['price_euro'].notna().mean(),
        'size_sqm': df['size_sqm'].notna().mean(),
        'city': df['city'].notna().mean()
    }

    for col, ratio in checks.items():
        if ratio < 0.8:
            logging.warning(f"[{col}] Low coverage: {ratio:.2f}")
        else:
            logging.info(f"[{col}] Good coverage: {ratio:.2f}")

def create_data_quality_report(df: pd.DataFrame, columns: list[str] = None, output_excel: str = None, output_html: str = None):
    report_df = calculate_data_quality(df, columns)

    if output_excel:
        export_report_to_excel(report_df, output_excel)
    
    if output_html:
        export_report_to_html(report_df, output_html)

    return report_df

def calculate_data_quality(df: pd.DataFrame, columns: list[str] = None):
    if columns is None:
        columns = df.columns
    else:
        columns = [col for col in columns if col in df.columns]
        
    total = len(df)
    
    report_df = pd.DataFrame({
        'column': columns,
        'unique_values': [df[col].nunique() for col in columns],
        'filled_%': [round(df[col].notna().sum() / total * 100, 2) for col in columns],
        'missing_%': [round(df[col].isna().sum() / total * 100, 2) for col in columns]
    })

    return report_df

def export_report_to_excel(report_df: pd.DataFrame, path: str):
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
    from openpyxl.styles import Border, Side, PatternFill
    
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        report_df.to_excel(writer, index=False, sheet_name='Report')
        ws = writer.sheets['Report']

        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 15
        
        max_row = ws.max_row

        red_fill = PatternFill(start_color='ffb3b3', end_color='ffb3b3', fill_type='solid')
        yellow_fill = PatternFill(start_color='ffffb3', end_color='ffffb3', fill_type='solid')
        green_fill = PatternFill(start_color='c1f0c1', end_color='c1f0c1', fill_type='solid')
        
        # red < 50
        ws.conditional_formatting.add(
            f"D2:D{max_row}",
            CellIsRule(operator='greaterThan', formula=['40'], fill=red_fill)
         )
        
        # yellow 50–80
        ws.conditional_formatting.add(
            f"D2:D{max_row}",
            CellIsRule(operator='between', formula=['40', '75'], fill=yellow_fill)
        )

        ws.conditional_formatting.add(
            f"D2:D{max_row}",
            CellIsRule(operator='lessThan', formula=['40'], fill=green_fill)
        )

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border

def export_report_to_html(report_df: pd.DataFrame, path: str):
    html = report_df.to_html(index=False, classes='table table-striped')
    with open(path, "w", encoding="utf-8") as f:
        f.write(f"""
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body {{ font-family: Arial; margin: 20px; }}
                table {{ border-collapse: collapse; width: 60%; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: center; }}
                th {{ background-color: #f2f2f2; }}
            </style>
        </head>
        <body>
            <h2>Data Quality Report</h2>
            {html}
        </body>
        </html>
        """)

    #return report_df
